#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
One-time bootstrap for accounts/users.xlsx — the editable login/estate-access
sheet CENDERA reads from (see app/auth.py).

Creates 2 admin accounts (see everything) plus one example estate-scoped
account per existing data/<estate>/ folder. Safe to re-run: if
accounts/users.xlsx already exists, it's left untouched (delete it yourself
first if you want a fresh bootstrap — your edits are never overwritten).

    cd "C:\\Documents\\oil_palm_dashboard\\oil_palm_dashboard"
    python scripts/setup_accounts.py
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ACCOUNTS_DIR = os.path.join(PROJECT_ROOT, "accounts")
OUTPUT_FILE = os.path.join(ACCOUNTS_DIR, "users.xlsx")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
ROLE_OPTIONS = ["admin", "user"]
COLUMNS = ["username", "password", "role", "estates"]


def discover_estates():
    if not os.path.isdir(DATA_DIR):
        return []
    return sorted(
        d for d in os.listdir(DATA_DIR) if os.path.isdir(os.path.join(DATA_DIR, d))
    )


def build_rows():
    rows = [
        ("admin1", "ChangeMe-Admin1", "admin", ""),
        ("admin2", "ChangeMe-Admin2", "admin", ""),
    ]
    for estate in discover_estates():
        slug = "".join(c.lower() if c.isalnum() else "_" for c in estate).strip("_")
        rows.append((f"{slug}_manager", f"ChangeMe-{estate}", "user", estate))
    return rows


def main():
    if os.path.exists(OUTPUT_FILE):
        sys.exit(
            f"{OUTPUT_FILE} already exists — leaving it alone so your edits "
            f"aren't lost. Delete it first if you want a fresh bootstrap."
        )
    os.makedirs(ACCOUNTS_DIR, exist_ok=True)

    rows = build_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    lists = wb.create_sheet("lists")
    for i, opt in enumerate(ROLE_OPTIONS, start=1):
        lists.cell(row=i, column=1, value=opt)
    lists.sheet_state = "hidden"

    fill = PatternFill("solid", fgColor="1F5C3A")
    hf = Font(bold=True, color="FFFFFF")
    for j, c in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill = fill
        cell.font = hf
        cell.alignment = Alignment(vertical="center")

    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    n = len(rows) + 1
    role_letter = get_column_letter(COLUMNS.index("role") + 1)
    dv = DataValidation(
        type="list", formula1=f"lists!$A$1:$A${len(ROLE_OPTIONS)}",
        allow_blank=False, showDropDown=False,
    )
    dv.error = "Pick 'admin' or 'user'."
    dv.errorTitle = "Invalid role"
    ws.add_data_validation(dv)
    dv.add(f"{role_letter}2:{role_letter}{n}")

    widths = {"username": 22, "password": 22, "role": 10, "estates": 30}
    for j, c in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 16)
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}")
    print(f"  {len(rows)} accounts: 2 admin + {len(rows) - 2} estate-scoped")
    print("  CHANGE THE PLACEHOLDER PASSWORDS before giving these out.")


if __name__ == "__main__":
    main()
