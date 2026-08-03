#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Precompute per-flight telemetry stats from the raw DJI flight-log exports.

The raw files (data/Flight Telemetry/<estate>/*.xlsx) run 90,000+ rows and
take 10-20+ seconds to parse — fine to run here, once, locally, but far too
slow to parse on every dashboard page load (and a real risk of timing out
the request on a slower/free-tier host). This script does the slow parse
once and writes a small summary (a few KB, one row per flight) that
core/flight_data.py reads instead — the live app never touches the raw
90k-row files at all.

Re-run this whenever you add/update a telemetry export.

    python scripts/build_flight_summary.py
"""

import os
import re
import glob
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TELEMETRY_DIR = os.path.join(PROJECT_ROOT, "data", "Flight Telemetry")
OUTPUT_FILE = os.path.join(PROJECT_ROOT, "outputs", "flight_telemetry_summary.xlsx")

_SRC_RE = re.compile(r"DJIFlightRecord_(\d{4})-(\d{2})-(\d{2})_\[(\d{2})-(\d{2})-(\d{2})\]")
COLUMNS = ["estate", "datetime_key", "duration_s", "height_avg_m", "height_max_m",
           "speed_avg_ms", "signal_avg_pct"]


def _source_file_key(name):
    m = _SRC_RE.search(str(name))
    return "".join(m.groups()) if m else None


def summarize_estate(estate_dir):
    files = sorted(glob.glob(os.path.join(estate_dir, "*.xlsx")))
    acc = {}
    for f in files:
        print(f"  reading {os.path.basename(f)} ...")
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        idx = {name: i for i, name in enumerate(header)}
        n_rows = 0
        for row in rows:
            key = _source_file_key(row[idx["Source_File"]])
            if key is None:
                continue
            t, h, sp, sigp = row[idx["Time_s"]], row[idx["Height_m"]], row[idx["Speed_ms"]], row[idx["Signal_Video"]]
            a = acc.setdefault(key, {"t_min": t, "t_max": t, "h_sum": 0.0, "h_max": h,
                                      "sp_sum": 0.0, "sig_sum": 0.0, "n": 0})
            a["t_min"] = min(a["t_min"], t)
            a["t_max"] = max(a["t_max"], t)
            a["h_sum"] += h
            a["h_max"] = max(a["h_max"], h)
            a["sp_sum"] += sp
            a["sig_sum"] += sigp
            a["n"] += 1
            n_rows += 1
        wb.close()
        print(f"    {n_rows} rows")
    return acc


def main():
    if not os.path.isdir(TELEMETRY_DIR):
        sys.exit(f"No {TELEMETRY_DIR} folder found.")
    estates = sorted(d for d in os.listdir(TELEMETRY_DIR) if os.path.isdir(os.path.join(TELEMETRY_DIR, d)))
    if not estates:
        sys.exit("No estate folders found under data/Flight Telemetry/.")

    all_rows = []
    for estate in estates:
        print(f"[{estate}]")
        acc = summarize_estate(os.path.join(TELEMETRY_DIR, estate))
        for key, a in acc.items():
            n = a["n"] or 1
            all_rows.append({
                "estate": estate, "datetime_key": key,
                "duration_s": round(a["t_max"] - a["t_min"], 1),
                "height_avg_m": round(a["h_sum"] / n, 2),
                "height_max_m": round(a["h_max"], 2),
                "speed_avg_ms": round(a["sp_sum"] / n, 3),
                "signal_avg_pct": round(a["sig_sum"] / n, 2),
            })
        print(f"  -> {len(acc)} flights")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "flight_telemetry_summary"
    fill = PatternFill("solid", fgColor="1F5C3A")
    hf = Font(bold=True, color="FFFFFF")
    for j, c in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill = fill
        cell.font = hf
    for i, row in enumerate(all_rows, start=2):
        for j, c in enumerate(COLUMNS, start=1):
            ws.cell(row=i, column=j, value=row[c])
    for j, c in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(c) + 2)
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"\nWrote {OUTPUT_FILE} ({len(all_rows)} flights total)")


if __name__ == "__main__":
    main()
