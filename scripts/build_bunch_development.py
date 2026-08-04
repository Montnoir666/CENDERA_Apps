#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bunch Development sync: consolidate per-flight bunch-detection workbooks into
one dashboard-ready file.

Estate-aware. Scans:
  data/Bunch Development/<Estate>/metadata/*_ALL_DETECTIONS.xlsx
      (one row per detected bunch: class, confidence, bbox, frame-level GPS)
  data/Bunch Development/<Estate>/annotated_frames/<flight>/<frame>.jpg
      (rendered frame with boxes drawn in, if it was exported)

Every real detection is kept, whether or not its frame image happens to be
available — frame_relpath is just left blank when there's no image to show.
No manual step here (unlike WC1): classification comes straight from the
detection model, so this is a full rebuild every run, not a merge.

A flight is dropped if its GPS doesn't actually match its claimed estate
(cross-checked against that estate's known location in the Agronomic
Assessment data) — better to leave a flight out than plot it in the wrong
place. See MAX_DISTANCE_KM below.

    python scripts/build_bunch_development.py
"""

import os
import re
import math
import glob
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_ROOT = os.path.join(BASE_DIR, "data", "Bunch Development")
OUTPUT_FILE = os.path.join(BASE_DIR, "outputs", "Bunch_development_synced.xlsx")
WC1_REFERENCE_FILE = os.path.join(BASE_DIR, "outputs", "WC1_agronomic_assessment_synced.xlsx")
IMAGE_EXTS = (".jpg", ".jpeg", ".png")
MAX_DISTANCE_KM = 10.0

CLASS_LABELS = {
    "Black_Bunch": "Black Bunch",
    "Black Bunch": "Black Bunch",
    "Ripe_Bunch": "Ripe Bunch",
    "Ripe Bunch": "Ripe Bunch",
}

_FOLDER_RE = re.compile(r"DJI_(\d{4})(\d{2})(\d{2})")


def date_from_folder(folder_name):
    m = _FOLDER_RE.match(str(folder_name))
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else "unknown"


def estate_reference_centers():
    """{estate: (lat, lng)} from the Agronomic Assessment data, used to sanity-
    check that a Bunch Development flight's GPS actually matches its estate."""
    if not os.path.exists(WC1_REFERENCE_FILE):
        return {}
    df = pd.read_excel(WC1_REFERENCE_FILE, sheet_name="assessment")
    df = df.dropna(subset=["latitude", "longitude"])
    df = df[(df["latitude"] != 0) & (df["longitude"] != 0)]
    return {est: (g["latitude"].mean(), g["longitude"].mean()) for est, g in df.groupby("estate")}


def km_distance(lat1, lng1, lat2, lng2):
    dlat = (lat2 - lat1) * 111.0
    dlng = (lng2 - lng1) * 111.0 * math.cos(math.radians((lat1 + lat2) / 2))
    return math.hypot(dlat, dlng)


def main():
    if not os.path.isdir(DATA_ROOT):
        sys.exit(f"No {DATA_ROOT} folder found.")
    estates = sorted(d for d in os.listdir(DATA_ROOT) if os.path.isdir(os.path.join(DATA_ROOT, d)))
    if not estates:
        sys.exit("No estate folders found under data/Bunch Development/.")

    ref_centers = estate_reference_centers()

    all_rows = []
    for estate in estates:
        meta_dir = os.path.join(DATA_ROOT, estate, "metadata")
        frames_dir = os.path.join(DATA_ROOT, estate, "annotated_frames")
        files = sorted(glob.glob(os.path.join(meta_dir, "*_ALL_DETECTIONS.xlsx")))
        if not files:
            print(f"[{estate}] no *_ALL_DETECTIONS.xlsx found in metadata/, skipping")
            continue
        ref = ref_centers.get(estate)

        with_image = 0
        for f in files:
            source_folder = os.path.basename(f).replace("_ALL_DETECTIONS.xlsx", "")
            df = pd.read_excel(f)

            if ref is not None:
                dist = km_distance(df["latitude"].mean(), df["longitude"].mean(), ref[0], ref[1])
                if dist > MAX_DISTANCE_KM:
                    print(f"[skip] {estate}/{source_folder}: GPS is {dist:.0f} km from "
                          f"{estate}'s known location — not actually this estate, excluding.")
                    continue

            df["estate"] = estate
            df["source_folder"] = source_folder
            df["assessment_date"] = date_from_folder(source_folder)
            df["detected_class"] = df["detected_class"].map(lambda c: CLASS_LABELS.get(str(c).strip(), str(c).strip()))

            def relpath(frame_filename):
                for ext in IMAGE_EXTS:
                    p = os.path.join(frames_dir, source_folder, frame_filename)
                    if os.path.exists(p):
                        return f"{estate}/annotated_frames/{source_folder}/{frame_filename}"
                return ""

            df["frame_relpath"] = df["frame_filename"].map(relpath)
            with_image += (df["frame_relpath"] != "").sum()
            print(f"[ok]  {estate}/{source_folder}: {len(df)} detections, "
                  f"{(df['frame_relpath'] != '').sum()} with a viewable frame")
            all_rows.append(df)

    if not all_rows:
        sys.exit("Nothing matched.")

    out = pd.concat(all_rows, ignore_index=True)
    cols = ["estate", "assessment_date", "source_folder", "detected_class", "confidence",
            "latitude", "longitude", "altitude_ft", "x_center_px", "y_center_px",
            "width_px", "height_px", "frame_filename", "video_filename", "frame_relpath"]
    out = out[[c for c in cols if c in out.columns]]

    print(f"\nTotal detections: {len(out)}")
    print(f"Estates: {sorted(out['estate'].unique())}")
    print(f"Dates:   {sorted(out['assessment_date'].unique())}")
    print(f"Classes: {sorted(out['detected_class'].unique())}")
    print(f"With a viewable frame: {(out['frame_relpath'] != '').sum()}")

    write_excel(out, OUTPUT_FILE)
    print(f"Wrote: {OUTPUT_FILE}")


def write_excel(df, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "detections"

    cols = list(df.columns)
    fill = PatternFill("solid", fgColor="1F5C3A")
    hf = Font(bold=True, color="FFFFFF")
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=str(c))
        cell.fill = fill
        cell.font = hf
        cell.alignment = Alignment(vertical="center")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, c in enumerate(cols, start=1):
            v = row[c]
            ws.cell(row=i, column=j, value="" if pd.isna(v) else v)

    for j, c in enumerate(cols, start=1):
        L = get_column_letter(j)
        w = max(12, min(42, len(str(c)) + 2))
        if c in ("source_folder", "frame_relpath", "video_filename"):
            w = 32
        ws.column_dimensions[L].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    main()
