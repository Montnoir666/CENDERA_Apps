#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WC1 sync: match retained frames (folders) with metadata (CSV) -> workbook.

Now estate-aware. It scans:  data/<Estate>/**/DJI_*   (flight folders)
falling back to a flat scan of the current folder if data/ has none.

Adds columns:
    estate            (from the data/<Estate>/ folder)
    assessment_month  (YYYY-MM, from the flight date in the folder name)
    frame_relpath     (path to the frame image, relative to this folder;
                       used by the dashboard to show the image)
plus the two you fill in:  agronomic_assessment_type (drop-down) + remarks.

Re-running MERGES: assessments you've already typed are preserved; new frames
come in blank.

    python --version                # 3.8+
    pip install pandas openpyxl
    python scripts/build_wc1_assessment.py
"""

import os
import re
import glob
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------------- #
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join("outputs", "WC1_agronomic_assessment_synced.xlsx")
IMAGE_EXTS = (".jpg", ".jpeg", ".png")
METADATA_SUFFIX = "_frame_metadata"
FRAME_COL_OVERRIDE = None
ASSESSMENT_OPTIONS = [
    "Nutrient Deficiency", "Pest & Disease",
    "Frond Pruning/Canopy Condition", "Ground Condition",
]
# --------------------------------------------------------------------------- #


def norm(name):
    if name is None:
        return ""
    base = os.path.basename(str(name)).strip().lower()
    return re.sub(r"\.(jpg|jpeg|png)$", "", base)


def month_from_folder(folder_name):
    m = re.search(r"DJI_(\d{4})(\d{2})(\d{2})", folder_name)
    return f"{m.group(1)}-{m.group(2)}" if m else "unknown"


def find_frame_column(cols):
    if FRAME_COL_OVERRIDE:
        return FRAME_COL_OVERRIDE
    lc = {c: str(c).lower() for c in cols}
    for c, l in lc.items():
        if "frame" in l and ("file" in l or "name" in l):
            return c
    for c, l in lc.items():
        if "frame" in l:
            return c
    for c, l in lc.items():
        if "file" in l or "name" in l:
            return c
    return None


def read_metadata(path):
    if path.lower().endswith(".csv"):
        return pd.read_csv(path)
    for eng in (None, "openpyxl"):
        try:
            return pd.read_excel(path, engine=eng) if eng else pd.read_excel(path)
        except Exception:
            continue
    raise RuntimeError(f"Could not read {path}")


def discover_flights(base):
    """Return list of (flight_name, flight_path, estate)."""
    flights = []
    data_root = os.path.join(base, "data")
    if os.path.isdir(data_root):
        for estate in sorted(os.listdir(data_root)):
            ep = os.path.join(data_root, estate)
            if not os.path.isdir(ep):
                continue
            for path in glob.glob(os.path.join(ep, "**", "DJI_*"), recursive=True):
                if os.path.isdir(path) and os.path.basename(path).upper().startswith("DJI_"):
                    flights.append((os.path.basename(path), path, estate))
    if not flights:  # fallback: flat scan
        for d in sorted(os.listdir(base)):
            p = os.path.join(base, d)
            if os.path.isdir(p) and d.upper().startswith("DJI_"):
                flights.append((d, p, "Unassigned"))
    return flights


def main():
    base = os.path.abspath(BASE_DIR)
    print(f"Base folder: {base}\n")

    flights = discover_flights(base)
    if not flights:
        sys.exit("No DJI_* flight folders found (neither under data/ nor flat).")

    all_rows = []
    frame_col_used = None

    for name, fpath, estate in sorted(flights, key=lambda x: (x[2], x[0])):
        imgs = []
        for ext in IMAGE_EXTS:
            imgs += glob.glob(os.path.join(fpath, f"*{ext}"))
        # de-dup (Windows glob is case-insensitive)
        img_map = {}
        for p in imgs:
            img_map.setdefault(norm(p), os.path.basename(p))
        img_keys = set(img_map)

        meta_path = None
        for ext in (".csv", ".xlsx", ".xls"):
            cand = fpath + METADATA_SUFFIX + ext
            if os.path.exists(cand):
                meta_path = cand
                break
        if meta_path is None:
            print(f"[skip] {estate}/{name}: no {name}{METADATA_SUFFIX}.csv")
            continue

        df = read_metadata(meta_path)
        fcol = find_frame_column(df.columns)
        if fcol is None:
            print(f"[skip] {estate}/{name}: frame column not found. Cols={list(df.columns)}")
            continue
        frame_col_used = fcol

        df["_key"] = df[fcol].map(norm)
        matched = df[df["_key"].isin(img_keys)].copy()

        rel_dir = os.path.relpath(fpath, base).replace("\\", "/")
        matched["frame_relpath"] = matched["_key"].map(
            lambda k: f"{rel_dir}/{img_map.get(k, '')}")
        matched.insert(0, "estate", estate)
        matched.insert(1, "assessment_month", month_from_folder(name))
        matched.insert(2, "source_folder", name)

        print(f"[ok]  {estate}/{name}: {len(img_keys)} imgs, {len(df)} rows -> "
              f"{len(matched)} matched")
        all_rows.append(matched)

    if not all_rows:
        sys.exit("Nothing matched. Check frame-filename column / FRAME_COL_OVERRIDE.")

    out = pd.concat(all_rows, ignore_index=True).drop(columns=["_key"])
    out["agronomic_assessment_type"] = ""
    out["remarks"] = ""

    # ---- merge previous assessments ----
    os.makedirs(os.path.dirname(os.path.join(base, OUTPUT_FILE)), exist_ok=True)
    existing = os.path.join(base, OUTPUT_FILE)
    if os.path.exists(existing):
        try:
            prev = pd.read_excel(existing, sheet_name="assessment")
            keys = ["source_folder", "frame_filename"]
            if set(keys).issubset(prev.columns):
                prev = prev.drop_duplicates(subset=keys).set_index(keys)
                kept = 0
                for i in out.index:
                    k = (out.at[i, "source_folder"], out.at[i, "frame_filename"])
                    if k in prev.index:
                        row = prev.loc[k]
                        at = row.get("agronomic_assessment_type", "")
                        rm = row.get("remarks", "")
                        if pd.notna(at) and str(at).strip():
                            out.at[i, "agronomic_assessment_type"] = at
                            kept += 1
                        if pd.notna(rm) and str(rm).strip():
                            out.at[i, "remarks"] = rm
                print(f"\nMerged {kept} existing assessments; {len(out)-kept} new/blank.")
        except Exception as e:
            print(f"[warn] merge skipped ({e}); writing fresh.")

    print(f"Total retained frames: {len(out)}")
    print(f"Estates: {sorted(out['estate'].unique())}")
    print(f"Months:  {sorted(out['assessment_month'].unique())}")
    print(f"Frame column used: '{frame_col_used}'")

    write_excel(out, os.path.join(base, OUTPUT_FILE))
    print(f"Wrote: {os.path.join(base, OUTPUT_FILE)}")


def write_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "assessment"
    lists = wb.create_sheet("lists")
    for i, opt in enumerate(ASSESSMENT_OPTIONS, start=1):
        lists.cell(row=i, column=1, value=opt)
    lists.sheet_state = "hidden"

    cols = list(df.columns)
    fill = PatternFill("solid", fgColor="1F5C3A")
    hf = Font(bold=True, color="FFFFFF")
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=str(c))
        cell.fill = fill
        cell.font = hf
        cell.alignment = Alignment(vertical="center")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, c in enumerate(cols, start=1):
            v = row[c]
            ws.cell(row=i, column=j, value="" if pd.isna(v) else v)

    n = len(df) + 1
    at_letter = get_column_letter(cols.index("agronomic_assessment_type") + 1)
    dv = DataValidation(type="list", formula1=f"lists!$A$1:$A${len(ASSESSMENT_OPTIONS)}",
                        allow_blank=True, showDropDown=False)
    dv.error = "Pick one of the 4 assessment types."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{at_letter}2:{at_letter}{n}")

    for j, c in enumerate(cols, start=1):
        L = get_column_letter(j)
        w = max(12, min(42, len(str(c)) + 2))
        if c in ("source_folder", "remarks", "agronomic_assessment_type", "frame_relpath"):
            w = 32
        ws.column_dimensions[L].width = w
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f"\nERROR: '{os.path.basename(path)}' is open in Excel. "
                 "Close it and re-run (your typed assessments are safe).")


if __name__ == "__main__":
    main()
